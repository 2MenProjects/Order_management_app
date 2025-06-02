from tkinter import *
from tkinter import ttk
import json
import os
from src.Model.NhanVien import NhanVien
from src.Model.TaiXe import TaiXe
from tkinter import messagebox as mb
from tkinter import filedialog as fd
import hashlib
import re
import openpyxl


class AccountPage:
    def __init__(self, root):
        self.root = root
        self.root.configure(bg="#f0f4f8")
        self.root.rowconfigure(1, weight=1)
        self.root.columnconfigure(0, weight=1)

        self.filename = "src/database/account.json"
        self.fileDriver = "src/database/driver.json"
        self.dstk = []
        self.dstx = []
        self.data = self.load_data_from_file(self.filename)
        self.dataShipper = self.load_data_from_file(self.fileDriver)

        self.dstk = [NhanVien.from_dict(user) for user in self.data]
        self.dstx = [TaiXe.from_dict(tx) for tx in self.dataShipper]

        self.treeview_bind_id = None

        self.create_frame_account_table()
        self.create_frame_duoi_table()

    def row_selection(self):
        selected_item = self.accountTable.focus()
        if selected_item:
            item_values = self.accountTable.item(selected_item)['values']
            if item_values:
                self.btn_remove.config(state="normal")
                self.btn_edit.config(state="normal")

                # Mở khóa các entry
                entries = [
                    self.entry_username, self.entry_fullname, self.entry_password,
                    self.entry_lcb, self.entry_phone, self.entry_address,
                    self.entry_cccd, self.entry_nganhang, self.entry_stk
                ]
                for entry in entries:
                    entry.config(state="normal")
                    entry.delete(0, END)

                self.entry_username.insert(0, item_values[0])
                self.entry_fullname.insert(0, item_values[1])
                self.entry_password.insert(0, item_values[2])
                self.entry_role.config(state="readonly")
                self.entry_role.set(item_values[3])

                luong = str(item_values[4]).replace(".", "").replace("đ", "")
                self.entry_lcb.insert(0, luong)
                self.entry_phone.insert(0, str(item_values[5]).zfill(10))
                self.entry_address.insert(0, item_values[6])
                self.entry_cccd.insert(0, str(item_values[7]).zfill(12))
                self.entry_nganhang.insert(0, item_values[8])
                self.entry_stk.insert(0, str(item_values[9]))

                # Đặt lại thành readonly
                for entry in entries:
                    entry.config(state="readonly")
                self.entry_role.config(state="disabled")

    def create_frame_account_table(self):
        def search():
            for item in self.accountTable.get_children():
                self.accountTable.delete(item)
            nd_timkiem = self.entry_tim_kiem.get().strip()
            self.user_timkiem = []
            for user in self.data:
                if nd_timkiem != "":
                    if str(user["tenTaiKhoan"]) == nd_timkiem or str(user["tenNguoiDung"]) == nd_timkiem or str(user["loaiTaiKhoan"]) == nd_timkiem or str(user["soDienThoai"]) == nd_timkiem or str(user["cccd"]) == nd_timkiem or str(user["nganHang"]) == nd_timkiem or str(user["stk_nganHang"]) == nd_timkiem:
                        self.user_timkiem.append(user)

                else:
                    self.user_timkiem.append(user)

            for index, item in enumerate(self.user_timkiem):
                tag = "evenrow" if index % 2 == 0 else "oddrow"
                self.accountTable.insert("", END, values=(
                    item["tenTaiKhoan"],
                    item["tenNguoiDung"],
                    item["matKhau"],
                    item["loaiTaiKhoan"],
                    "{:,}đ".format(int(item["luongCoBan"])).replace(",", "."),
                    str(item["soDienThoai"]),
                    str(item["diaChi"]),
                    str(item["cccd"]),
                    item.get("nganHang", ""),
                    str(item["stk_nganHang"])), tags=(tag,))

        self.frame_search = Frame(self.root)

        self.frame_search.grid(row=0, column=0, sticky="e", padx=50, pady=5)

        self.entry_tim_kiem = Entry(
            self.frame_search, width=20, font=("Arial", 13))
        self.entry_tim_kiem.grid(row=0, column=0, padx=(0, 20))

        self.btn_search = Button(
            self.frame_search, width=10, text="Tìm kiếm", font=("Arial", 13, "bold"),
            bg="black", fg="white", bd=1, command=search
        )
        self.btn_search.grid(row=0, column=1)

        self.table_frame = Frame(self.root)
        self.table_frame.grid(row=1, column=0, sticky="news")

        self.table_frame.rowconfigure(0, weight=1)
        self.table_frame.columnconfigure(0, weight=1)

        columns = ("#1", "#2", "#3", "#4", "#5", "#6", "#7", "#8", "#9", "#10")
        self.accountTable = ttk.Treeview(
            self.table_frame, columns=columns, show="headings", height=7)
        self.accountTable.grid(row=0, column=0, sticky="nsew")

        self.accountTable.heading("#1", text="Tên tài khoản")
        self.accountTable.heading("#2", text="Tên người dùng")
        self.accountTable.heading("#3", text="Mật khẩu")
        self.accountTable.heading("#4", text="Loại tài khoản")
        self.accountTable.heading("#5", text="Lương cơ bản")
        self.accountTable.heading("#6", text="SĐT")
        self.accountTable.heading("#7", text="Địa chỉ")
        self.accountTable.heading("#8", text="CCCD")
        self.accountTable.heading("#9", text="Ngân hàng")
        self.accountTable.heading("#10", text="STK")

        self.accountTable.tag_configure("oddrow", background="#f0f0f0")
        self.accountTable.tag_configure("evenrow", background="#ffffff")

        for i in range(1, 11):
            self.accountTable.column(f"#{i}", width=80, anchor="center")

        self.scrollbar = ttk.Scrollbar(
            self.table_frame, orient="vertical", command=self.accountTable.yview)
        self.accountTable.configure(yscroll=self.scrollbar.set)
        # self.scrollbar.pack(side="right", fill="y")
        # self.accountTable.pack(fill="both", expand=True)

        self.scrollbar.grid(row=0, column=1, sticky="ns")

        for index, item in enumerate(self.data):
            tag = "evenrow" if index % 2 == 0 else "oddrow"
            self.accountTable.insert("", END, values=(
                item["tenTaiKhoan"],
                item["tenNguoiDung"],
                item["matKhau"],
                item["loaiTaiKhoan"],
                "{:,}đ".format(int(item["luongCoBan"])).replace(",", "."),
                str(item["soDienThoai"]),
                str(item["diaChi"]),
                str(item["cccd"]),
                item.get("nganHang", ""),
                str(item["stk_nganHang"])), tags=(tag,))

        self.treeview_bind_id = self.accountTable.bind(
            "<<TreeviewSelect>>", lambda e: self.row_selection())

    def create_frame_duoi_table(self):
        self.frame_duoi_table = Frame(self.root)
        self.frame_duoi_table.grid(row=2, column=0, sticky="news", pady=10)

        self.frame_duoi_table.columnconfigure(0, weight=3)
        self.frame_duoi_table.columnconfigure(1, weight=1)
        self.frame_duoi_table.rowconfigure(0, weight=1)

        self.frame_left = Frame(self.frame_duoi_table)
        self.frame_left.grid(row=0, column=0, sticky="nsew")

        self.frame_left.columnconfigure(0, weight=1)
        self.frame_left.rowconfigure(0, weight=0)
        self.frame_left.rowconfigure(1, weight=1)

        self.create_frame_button()
        self.create_entry()

        self.create_frame_hien_thi_so_luong()

    def create_frame_button(self):
        def refresh_table():
            for item in self.accountTable.get_children():
                self.accountTable.delete(item)

            for index, user in enumerate(self.dstk):
                tag = "evenrow" if index % 2 == 0 else "oddrow"
                self.accountTable.insert("", END, values=(
                    user.ten_tai_khoan, user.ten_nguoi_dung,
                    user.mat_khau, user.loai_tai_khoan,
                    "{:,}đ".format(user.luong_co_ban).replace(",", "."),
                    str(user.so_dien_thoai),
                    str(user.dia_chi),
                    str(user.cccd), user.ngan_hang, str(user.stk_ngan_hang)), tags=(tag,))

                self.lbl_so_luong_taikhoan.config(
                    text=f"Tổng số tài khoản: {len(self.dstk)}")

        def refresh_entry():
            self.entry_username.config(state="normal")
            self.entry_fullname.config(state="normal")
            self.entry_password.config(state="normal")
            self.entry_lcb.config(state="normal")
            self.entry_address.config(state="normal")
            self.entry_phone.config(state="normal")
            self.entry_cccd.config(state="normal")
            self.entry_nganhang.config(state="normal")
            self.entry_stk.config(state="normal")
            self.entry_role.config(state="readonly")

            self.entry_username.delete(0, END)
            self.entry_password.delete(0, END)
            self.entry_fullname.delete(0, END)
            self.entry_lcb.delete(0, END)
            self.entry_address.delete(0, END)
            self.entry_phone.delete(0, END)
            self.entry_cccd.delete(0, END)
            self.entry_nganhang.delete(0, END)
            self.entry_stk.delete(0, END)
            self.entry_role.set("user")

            self.entry_username.config(state="readonly")
            self.entry_password.config(state="readonly")
            self.entry_fullname.config(state="readonly")
            self.entry_lcb.config(state="readonly")
            self.entry_address.config(state="readonly")
            self.entry_phone.config(state="readonly")
            self.entry_cccd.config(state="readonly")
            self.entry_nganhang.config(state="readonly")
            self.entry_stk.config(state="readonly")
            self.entry_role.config(state="disabled")

        def add_user():
            if self.btn_add["text"] == "Thêm":
                self.entry_username.config(state="normal")
                self.entry_fullname.config(state="normal")
                self.entry_password.config(state="normal")
                self.entry_lcb.config(state="normal")
                self.entry_address.config(state="normal")
                self.entry_phone.config(state="normal")
                self.entry_cccd.config(state="normal")
                self.entry_nganhang.config(state="normal")
                self.entry_stk.config(state="normal")

                self.entry_username.delete(0, END)
                self.entry_fullname.delete(0, END)
                self.entry_password.delete(0, END)
                self.entry_lcb.delete(0, END)
                self.entry_address.delete(0, END)
                self.entry_phone.delete(0, END)
                self.entry_cccd.delete(0, END)
                self.entry_nganhang.delete(0, END)
                self.entry_stk.delete(0, END)
                self.entry_role.set("Nhân viên giao dịch")

                self.entry_role.config(state="readonly")

                self.btn_add.config(text="Xác nhận")
                self.btn_remove.config(state="disabled")
                self.btn_edit.config(state="disabled")
                self.btn_refresh.config(state="disabled")
                self.btn_cancel.config(state="normal")

                self.treeview_bind_id = self.accountTable.unbind(
                    "<<TreeviewSelect>>")

            elif self.btn_add["text"] == "Xác nhận":
                username = self.entry_username.get()
                password = self.entry_password.get()
                fullname = self.entry_fullname.get()
                role = self.entry_role.get()
                sdt = self.entry_phone.get().strip()
                diaChi = self.entry_address.get().strip()
                cccd = self.entry_cccd.get().strip()
                nganHang = self.entry_nganhang.get().strip()
                stk = self.entry_stk.get().strip()
                lcb_raw = self.entry_lcb.get().strip()

                if username == "":
                    mb.showwarning("Lỗi", "Vui lòng nhập username!")
                    return

                for user in self.dstk:
                    if user.ten_tai_khoan == username:
                        mb.showerror(
                            "Lỗi", "Tài khoản này đã tồn tại. Vui lòng nhập tên tài khoản khác!")
                        return

                if fullname == "":
                    mb.showwarning("Lỗi", "Vui lòng nhập họ tên!")
                    return

                if password == "":
                    mb.showwarning("Lỗi", "Vui lòng nhập mật khẩu!")
                    return

                if sdt == "":
                    mb.showwarning("Lỗi", "Vui lòng nhập số điện thoại!")
                    return

                if len(sdt) != 10:
                    mb.showwarning(
                        "Lỗi", "Vui lòng nhập đúng định dạng số điện thoại (10 chữ số)!")
                    return

                if cccd == "":
                    mb.showwarning("Lỗi", "Vui lòng nhập CCCD!")
                    return

                if len(cccd) != 12:
                    mb.showwarning(
                        "Lỗi", "Vui lòng nhập đúng định dạng CCCD (12 chữ số)!")
                    return

                try:
                    lcb = int(lcb_raw) if lcb_raw != "" else 0
                except ValueError:
                    mb.showwarning(
                        "Lỗi", "Lương cơ bản không hợp lệ! Vui lòng nhập số.")
                    return

                for user in self.dstk:
                    if cccd == user.cccd:
                        mb.showwarning(
                            "Lỗi", "Đã tồn tại nhân viên có CCCD này!")
                        return

                for tx in self.dstx:
                    if cccd == tx.cccd:
                        mb.showwarning(
                            "Lỗi", "Đã tồn tại nhân viên có CCCD này!")
                        return

                for user in self.dstk:
                    if sdt == user.so_dien_thoai:
                        mb.showwarning(
                            "Lỗi", "Đã tồn tại nhân viên có số điện thoại này!")
                        return

                for tx in self.dstx:
                    if sdt == tx.soDienThoai:
                        mb.showwarning(
                            "Lỗi", "Đã tồn tại nhân viên có số điện thoại này!")
                        return

                user = NhanVien(username, fullname, password, role, sdt,
                                str(diaChi), cccd, nganHang, stk, int(lcb))
                self.dstk.append(user)
                self.save_data(self.filename)
                mb.showinfo(
                    "Thông báo", f"Thêm tài khoản {username} thành công")

                refresh_table()
                refresh_entry()
                self.btn_add.config(text="Thêm")
                self.btn_refresh.config(state="normal")
                self.btn_cancel.config(state="disabled")

                self.accountTable.bind(
                    "<<TreeviewSelect>>", lambda e: self.row_selection())

        def remove_user():
            username = self.entry_username.get()
            answer = mb.askyesno(
                "Thông báo", f"Bạn có chắc chắn muốn xóa tài khoản {username} không?")
            if answer:
                for user in self.dstk:
                    if user.ten_tai_khoan == username:
                        self.dstk.remove(user)
                        self.save_data(self.filename)
                        refresh_table()
                        refresh_entry()
                        mb.showinfo(
                            "Thông báo", f"Đã xóa thành công tài khoản {username}.")

                        self.btn_remove.config(state="disabled")
                        self.btn_edit.config(state="disabled")
                        return
            return

        def edit():
            username = self.entry_username.get()
            old_pass = self.entry_password.get()
            old_cccd = self.entry_cccd.get()
            old_sdt = self.entry_phone.get()

            if self.btn_edit["text"] == "Sửa":
                self.btn_add.config(state="disabled")
                self.btn_refresh.config(state="disabled")
                self.btn_remove.config(state="disabled")

                self.btn_edit.config(text="Xác nhận")
                self.btn_cancel.config(state="normal")

                self.entry_fullname.config(state="normal")
                self.entry_password.config(state="normal")
                self.entry_lcb.config(state="normal")
                self.entry_address.config(state="normal")
                self.entry_cccd.config(state="normal")
                self.entry_nganhang.config(state="normal")
                self.entry_stk.config(state="normal")
                self.entry_phone.config(state="normal")
                self.entry_role.config(state="readonly")

                self.treeview_bind_id = self.accountTable.unbind(
                    "<<TreeviewSelect>>")

            else:
                new_fullname = self.entry_fullname.get().strip()
                new_password = self.entry_password.get().strip()
                new_lcb_raw = self.entry_lcb.get().strip()
                new_address = self.entry_address.get().strip()
                new_phone = self.entry_phone.get().strip()
                new_cccd = self.entry_cccd.get().strip()
                new_nganhang = self.entry_nganhang.get().strip()
                new_stk = self.entry_stk.get().strip()
                new_role = self.entry_role.get()

                if not new_fullname:
                    return mb.showwarning("Lỗi", "Vui lòng nhập họ tên!")

                if not new_password:
                    return mb.showwarning("Lỗi", "Vui lòng nhập mật khẩu!")

                if not new_phone:
                    return mb.showwarning("Lỗi", "Vui lòng nhập số điện thoại!")

                if not new_cccd:
                    return mb.showwarning("Lỗi", "Vui lòng nhập CCCD!")

                if len(new_cccd) != 12:
                    mb.showwarning(
                        "Lỗi", "Vui lòng nhập đúng định dạng CCCD (12 chữ số)!")
                    return

                for user in self.dstk:
                    if user.ten_tai_khoan != username and user.cccd == new_cccd:
                        return mb.showwarning("Lỗi", "Đã tồn tại nhân viên có CCCD này!")

                for tx in self.dstx:
                    if tx.maTaiXe != username and tx.cccd == new_cccd:
                        return mb.showwarning("Lỗi", "Đã tồn tại nhân viên có CCCD này!")

                if len(new_phone) != 10:
                    mb.showwarning(
                        "Lỗi", "Vui lòng nhập đúng định dạng số điện thoại (10 chữ số)!")
                    return

                for user in self.dstk:
                    if user.ten_tai_khoan != username and user.so_dien_thoai == new_phone:
                        return mb.showwarning("Lỗi", "Đã tồn tại nhân viên có số điện thoại này!")
                for tx in self.dstx:
                    if tx.maTaiXe != username and tx.soDienThoai == new_phone:
                        return mb.showwarning("Lỗi", "Đã tồn tại nhân viên có số điện thoại này!")

                try:
                    new_lcb = int(new_lcb_raw) if new_lcb_raw != "" else 0
                except ValueError:
                    mb.showwarning(
                        "Lỗi", "Lương cơ bản không hợp lệ! Vui lòng nhập số.")
                    return

                answer = mb.askyesno(
                    "Thông báo", f"Bạn có chắc muốn đổi thông tin cho tài khoản {username} không?")
                if answer:
                    for user in self.dstk:
                        if user.ten_tai_khoan == username:
                            def hash_password(password):
                                return hashlib.sha256(password.encode("utf-8")).hexdigest()

                            user.ten_nguoi_dung = new_fullname
                            if old_pass == new_password:
                                user.mat_khau = new_password
                            else:
                                user.mat_khau = hash_password(new_password)
                            user.loai_tai_khoan = new_role
                            user.luong_co_ban = int(new_lcb)
                            user.so_dien_thoai = new_phone
                            user.dia_chi = new_address
                            user.cccd = new_cccd
                            user.ngan_hang = new_nganhang
                            user.stk_ngan_hang = new_stk

                            self.save_data(self.filename)
                            refresh_table()
                            refresh_entry()
                            mb.showinfo(
                                "Thông báo", f"Đã thay đổi thông tin thành công cho tài khoản {username}")

                            self.btn_add.config(state="normal")
                            self.btn_refresh.config(state="normal")
                            self.btn_edit.config(text="Sửa")
                            self.btn_remove.config(state="disabled")
                            self.btn_edit.config(state="disabled")
                            self.btn_cancel.config(state="disabled")

                            self.accountTable.bind(
                                "<<TreeviewSelect>>", lambda e: self.row_selection())

                            return

        def cancel():
            if self.btn_add["text"] == "Xác nhận":
                answer = mb.askyesno(
                    "Thông báo", "Bạn có chắc muốn hủy thao tác thêm không?")

                if answer:
                    self.btn_add.config(text="Thêm")
                    self.btn_refresh.config(state="normal")
                    self.btn_cancel.config(state="disabled")

                    refresh_entry()

                    self.accountTable.bind(
                        "<<TreeviewSelect>>", lambda e: self.row_selection())
                    return
                return

            elif self.btn_edit["text"] == "Xác nhận":
                answer = mb.askyesno(
                    "Thông báo", "Bạn có chắc muốn hủy thao tác sửa không?")

                if answer:
                    self.btn_edit.config(text="Sửa")
                    self.btn_edit.config(state="disabled")
                    self.btn_add.config(state="normal")
                    self.btn_remove.config(state="disabled")
                    self.btn_refresh.config(state="normal")
                    self.btn_cancel.config(state="disabled")
                    refresh_entry()

                    self.accountTable.bind(
                        "<<TreeviewSelect>>", lambda e: self.row_selection())

                    return
                return

        def xuat_excel():
            if not self.dstk:
                mb.showwarning("Thông báo", "Không có dữ liệu để xuất!")
                return

            file_path = fd.asksaveasfilename(defaultextension=".xlsx", filetypes=[
                                             ("Excel files", "*.xlsx")], initialfile="Danh_sach_tai_khoan.xlsx")
            if not file_path:
                return

            wb = openpyxl.Workbook()
            ws = wb.active  # Worksheet
            ws.title = "Danh sách nhân viên giao dịch"

            headers = ["Tên tài khoản", "Tên người dùng", "Chức vụ",
                       "Số điện thoại", "Địa chỉ", "CCCD", "Ngân hàng", "STK ngân hàng", "Lương cơ bản"]
            ws.append(headers)

            for nv in self.dstk:
                row = [nv.ten_tai_khoan, nv.ten_nguoi_dung, nv.loai_tai_khoan, str(nv.so_dien_thoai), nv.dia_chi, str(
                    nv.cccd), nv.ngan_hang, str(nv.stk_ngan_hang), "{:,}đ".format(int(nv.luong_co_ban)).replace(",", ".")]

                ws.append(row)

            try:
                wb.save(file_path)
                mb.showinfo("Thành công", "Xuất file Excel thành công!")
            except Exception as e:
                mb.showerror("Lỗi", f"Không thể lưu file:\n{e}")

        def nhap_excel():
            file_path = fd.askopenfilename(
                filetypes=[("Excel files", "*.xlsx")])
            if not file_path:
                return

            try:
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active

                errors = []

                # bỏ dòng tiêu đề, chỉ lấy từ dòng 2 trở đi.
                rows = list(ws.iter_rows(min_row=2, values_only=True))

                for idx, row in enumerate(rows, start=2):
                    ma, ten, chuc_vu, sdt, diaChi, cccd, nganHang, stk, lcb = [
                        str(cell or "") for cell in row]

                    luong = lcb.replace("đ", "").replace(
                        ".", "") if lcb else "0"

                    row_errors = []

                    # Kiểm tra định dạng SĐT
                    if not sdt.isdigit() or len(sdt) != 10:
                        row_errors.append(
                            f"SĐT '{sdt}' không hợp lệ (phải đủ 10 chữ số).")

                    if not cccd.isdigit() or len(cccd) != 12:
                        row_errors.append(
                            f"CCCD '{cccd}' không hợp lệ (phải đủ 12 chữ số).")

                    # Kiểm tra trùng CCCD
                    cccd_conflict = any(
                        (tx.cccd == cccd and tx.maTaiXe != ma) for tx in self.dstx
                    ) or any(
                        (user.cccd == cccd and user.ten_tai_khoan != ma) for user in self.dstk
                    )

                    if cccd_conflict:
                        row_errors.append(
                            f"CCCD '{cccd}' đã tồn tại ở nhân viên khác.")

                    # Kiểm tra trùng SĐT
                    sdt_conflict = any(
                        (tx.soDienThoai == sdt and tx.maTaiXe != ma) for tx in self.dstx
                    ) or any(
                        (user.so_dien_thoai == sdt and user.ten_tai_khoan != ma) for user in self.dstk
                    )

                    if sdt_conflict:
                        row_errors.append(
                            f"SĐT '{sdt}' đã tồn tại ở nhân viên khác.")

                    if row_errors:
                        error_message = f"- Dòng {idx}: " + \
                            "; ".join(row_errors)
                        errors.append(error_message)
                        continue

                    mat_khau = "1"

                    for nv in self.dstk:
                        if nv.ten_tai_khoan == ma:
                            # giữ lại mật khẩu cũ
                            mat_khau = nv.mat_khau
                            break

                    nv_moi = NhanVien(
                        ten_tai_khoan=ma,
                        ten_nguoi_dung=ten,
                        mat_khau=mat_khau,  # mật khẩu mặc định
                        loai_tai_khoan=chuc_vu,
                        so_dien_thoai=sdt,
                        dia_chi=str(diaChi),
                        cccd=cccd,
                        ngan_hang=nganHang,
                        stk_ngan_hang=stk,
                        luong_co_ban=int(luong),
                        is_hashed=True
                    )

                    found = False
                    for i, nv in enumerate(self.dstk):
                        if nv.ten_tai_khoan == ma:
                            self.dstk[i] = nv_moi  # cập nhật nhân viên
                            found = True
                            break

                    if not found:
                        self.dstk.append(nv_moi)

                self.save_data(self.filename)
                refresh_table()

                if errors:
                    mb.showwarning(
                        "Một số dòng bị lỗi",
                        "\n".join(errors)
                    )
                else:
                    mb.showinfo(
                        "Thành công", "Nhập dữ liệu từ Excel thành công!")

            except Exception as e:
                mb.showerror("Lỗi", f"Không thể đọc file:\n{e}")

        self.frame_buttons = Frame(self.frame_left)
        self.frame_buttons.grid(row=0, column=0, sticky="news", pady=10)

        self.btn_add = Button(self.frame_buttons, text="Thêm",
                              width=15, height=2, font=("Arial", 13, "bold"), bg="blue", fg="white", command=add_user)
        self.btn_add.grid(row=0, column=0, padx=10)

        self.btn_remove = Button(self.frame_buttons, text="Xóa",
                                 width=15, height=2, font=("Arial", 13, "bold"), bg="#009df7", fg="white", state="disabled", command=remove_user)
        self.btn_remove.grid(row=0, column=1, padx=10)

        self.btn_edit = Button(self.frame_buttons, text="Sửa",
                               width=15, height=2, font=("Arial", 13, "bold"), bg="#00f7f7", fg="white", state="disabled", command=edit)
        self.btn_edit.grid(row=0, column=2, padx=10)

        self.btn_refresh = Button(self.frame_buttons, text="Làm mới", width=15, height=2, font=(
            "Arial", 13, "bold"), bg="#4287f5", fg="white", command=refresh_table)
        self.btn_refresh.grid(row=1, column=0, padx=10, pady=5)

        self.btn_cancel = Button(self.frame_buttons, text="Hủy", width=15, height=2, font=(
            "Arial", 13, "bold"), bg="#f70021", fg="white", state="disabled", command=cancel)
        self.btn_cancel.grid(row=1, column=1, padx=10, pady=5)

        self.btn_xuat_excel = Button(self.frame_buttons, text="Xuất file Excel", width=15, height=2, font=(
            "Arial", 13, "bold"), bg="#00B050", fg="white", state="normal", command=xuat_excel)
        self.btn_xuat_excel.grid(row=1, column=2, padx=10, pady=5)

        self.btn_nhap_excel = Button(self.frame_buttons, text="Nhập file Excel", width=15, height=2, font=(
            "Arial", 13, "bold"), bg="#00B050", fg="white", state="normal", command=nhap_excel)
        self.btn_nhap_excel.grid(row=0, column=3, padx=10, pady=5)

    def create_frame_hien_thi_so_luong(self):
        self.frame_sl_taikhoan = Frame(self.frame_duoi_table)
        self.frame_sl_taikhoan.grid(row=0, column=1, sticky="ensw", padx=20)

        tong_so_tk = len(self.dstk)
        self.lbl_so_luong_taikhoan = Label(
            self.frame_sl_taikhoan, text=f"Tổng số tài khoản: {tong_so_tk}", font=("Arial", 12), anchor="e")
        self.lbl_so_luong_taikhoan.grid(row=0, column=0, sticky="ew")

    def create_entry(self):
        self.frame_entries = Frame(self.frame_left)
        self.frame_entries.grid(row=1, column=0, sticky="news", pady=10)

        def validate_max_length(max_len):
            def validator(value):
                if value == "":
                    return True
                return bool(re.fullmatch(r"\d{0," + str(max_len) + r"}", value))
            return validator

        vcmd_sdt = self.root.register(validate_max_length(10))
        vcmd_cccd = self.root.register(validate_max_length(12))
        vcmd_stk = self.root.register(validate_max_length(20))
        vcmd_lcb = self.root.register(validate_max_length(10))

        # Username
        self.lbl_username = Label(
            self.frame_entries, text="Tên tài khoản: ", width=15, font=("Arial", 13))
        self.lbl_username.grid(row=0, column=0, padx=10, pady=8)
        self.entry_username = Entry(
            self.frame_entries,
            width=20,
            font=("Arial", 13),
            bd=1,
            relief="solid",
            highlightbackground="gray",
            highlightthickness=2,
            highlightcolor="blue",
            state="readonly")  # ko cho người dùng gõ vào
        self.entry_username.grid(row=0, column=1, padx=10, pady=8)

        # Tên người dùng
        self.lbl_fullname = Label(
            self.frame_entries, text="Tên người dùng: ", width=15, font=("Arial", 13))
        self.lbl_fullname.grid(row=1, column=0, padx=10, pady=8)
        self.entry_fullname = Entry(
            self.frame_entries,
            width=20,
            font=("Arial", 13),
            bd=1,
            relief="solid",
            highlightbackground="gray",
            highlightthickness=2,
            highlightcolor="blue",
            state="readonly")
        self.entry_fullname.grid(row=1, column=1, padx=10, pady=8)

        # Mật khẩu
        self.lbl_password = Label(
            self.frame_entries, text="Mật khẩu: ", width=15, font=("Arial", 13))
        self.lbl_password.grid(row=2, column=0, padx=10, pady=8)
        self.entry_password = Entry(
            self.frame_entries,
            width=20,
            font=("Arial", 13),
            bd=1,
            relief="solid",
            highlightbackground="gray",
            highlightthickness=2,
            highlightcolor="blue",
            state="readonly")
        self.entry_password.grid(row=2, column=1, padx=10, pady=8)

        # Role
        self.lbl_role = Label(
            self.frame_entries, text="Loại tài khoản: ", width=15, font=("Arial", 13))
        self.lbl_role.grid(row=3, column=0, padx=10, pady=8)
        self.entry_role = ttk.Combobox(
            self.frame_entries,
            width=20,
            font=("Arial", 13),
            state="disabled",  # ko cho người dùng bấm vào
            values=["Quản lý", "Nhân viên giao dịch"])
        self.entry_role.current(1)  # mặc định là user
        self.entry_role.grid(row=3, column=1, padx=10, pady=8)

        # Lương cơ bản
        self.lbl_lcb = Label(
            self.frame_entries, text="Lương cơ bản: ", width=15, font=("Arial", 13))
        self.lbl_lcb.grid(row=4, column=0, padx=10, pady=8)
        self.entry_lcb = Entry(
            self.frame_entries,
            width=20,
            font=("Arial", 13),
            bd=1,
            relief="solid",
            highlightbackground="gray",
            highlightthickness=2,
            highlightcolor="blue",
            state="readonly",
            validate="key",
            validatecommand=(vcmd_lcb, "%P"))
        self.entry_lcb.grid(row=4, column=1, padx=10, pady=8)

        # SĐT
        self.lbl_phone = Label(
            self.frame_entries, text="Số điện thoại: ", width=15, font=("Arial", 13))
        self.lbl_phone.grid(row=0, column=2, padx=10, pady=8)
        self.entry_phone = Entry(
            self.frame_entries,
            width=20,
            font=("Arial", 13),
            bd=1,
            relief="solid",
            highlightbackground="gray",
            highlightthickness=2,
            highlightcolor="blue",
            state="readonly",
            validate="key",
            validatecommand=(vcmd_sdt, "%P"))
        self.entry_phone.grid(row=0, column=3, padx=10, pady=8)

        # Địa chỉ
        self.lbl_address = Label(
            self.frame_entries, text="Địa chỉ: ", width=15, font=("Arial", 13))
        self.lbl_address.grid(row=1, column=2, padx=10, pady=8)
        self.entry_address = Entry(
            self.frame_entries,
            width=20,
            font=("Arial", 13),
            bd=1,
            relief="solid",
            highlightbackground="gray",
            highlightthickness=2,
            highlightcolor="blue",
            state="readonly")
        self.entry_address.grid(row=1, column=3, padx=10, pady=8)

        # CCCD
        self.lbl_cccd = Label(
            self.frame_entries, text="CCCD: ", width=15, font=("Arial", 13))
        self.lbl_cccd.grid(row=2, column=2, padx=10, pady=8)
        self.entry_cccd = Entry(
            self.frame_entries,
            width=20,
            font=("Arial", 13),
            bd=1,
            relief="solid",
            highlightbackground="gray",
            highlightthickness=2,
            highlightcolor="blue",
            state="readonly",
            validate="key",
            validatecommand=(vcmd_cccd, "%P"))
        self.entry_cccd.grid(row=2, column=3, padx=10, pady=8)

        # Ngân hàng
        self.lbl_nganhang = Label(
            self.frame_entries, text="Ngân hàng: ", width=15, font=("Arial", 13))
        self.lbl_nganhang.grid(row=3, column=2, padx=10, pady=8)
        self.entry_nganhang = Entry(
            self.frame_entries,
            width=20,
            font=("Arial", 13),
            bd=1,
            relief="solid",
            highlightbackground="gray",
            highlightthickness=2,
            highlightcolor="blue",
            state="readonly")
        self.entry_nganhang.grid(row=3, column=3, padx=10, pady=8)

        # Ngân hàng
        self.lbl_stk = Label(
            self.frame_entries, text="Số tài khoản: ", width=15, font=("Arial", 13))
        self.lbl_stk.grid(row=4, column=2, padx=10, pady=8)
        self.entry_stk = Entry(
            self.frame_entries,
            width=20,
            font=("Arial", 13),
            bd=1,
            relief="solid",
            highlightbackground="gray",
            highlightthickness=2,
            highlightcolor="blue",
            state="readonly",
            validate="key",
            validatecommand=(vcmd_stk, "%P"))
        self.entry_stk.grid(row=4, column=3, padx=10, pady=8)

    def load_data_from_file(self, filename):
        try:
            with open(filename, "r", encoding="utf-8") as file:
                return json.load(file)
        except (json.JSONDecodeError, FileNotFoundError):
            return []

    def save_data(self, filename):
        data = [user.to_dict() for user in self.dstk]
        with open(filename, "w", encoding="utf-8") as file:
            json.dump(data, file, ensure_ascii=False, indent=4)
