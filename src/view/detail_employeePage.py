from tkinter import *
from tkinter import messagebox as mb
from tkinter import ttk
from tkinter import filedialog as fd
import json
from src.Model.TaiXe import TaiXe
from src.Model.LichLamViec import LichLamViec
import random
import re
import openpyxl
from datetime import datetime, date
import calendar


class Detail_EmployeePage:
    def __init__(self, root, switch_page, maNhanVien):
        self.root = root
        self.root.configure(bg="#f0f4f8")
        self.root.columnconfigure(0, weight=1)

        self.switch_page = switch_page

        self.maNhanVien = maNhanVien
        self.filename = "src/database/detail_employee.json"
        self.file_driver = "src/database/driver.json"
        self.fileyear = "src/database/years.json"
        self.file_account = "src/database/account.json"

        self.dataLich = self.load_detail_employee(self.filename)

        hoten, luongCoBan, chucVu = self.get_ten_nhan_vien_va_luong_co_ban_va_chuc_vu(
            maNhanVien)

        frame_title = Frame(self.root, bg="#f0f4f8")
        frame_title.grid(row=0, column=0, sticky="w", padx=10, pady=10)

        Label(frame_title, text="Mã nhân viên: ", font=(
            "Arial", 13, "bold"), bg="#f0f4f8").pack(side=LEFT)
        Label(frame_title, text=maNhanVien, font=(
            "Arial", 13), bg="#f0f4f8").pack(side=LEFT)

        Label(frame_title, text=" - Họ tên: ",
              font=("Arial", 13, "bold"), bg="#f0f4f8").pack(side=LEFT)
        Label(frame_title, text=hoten, font=(
            "Arial", 13), bg="#f0f4f8").pack(side=LEFT)

        Label(frame_title, text=" - Lương cơ bản: ",
              font=("Arial", 13, "bold"), bg="#f0f4f8").pack(side=LEFT)
        Label(frame_title, text=f"{luongCoBan}/h",
              font=("Arial", 13), bg="#f0f4f8").pack(side=LEFT)

        Label(frame_title, text=" - Chức vụ: ",
              font=("Arial", 13, "bold"), bg="#f0f4f8").pack(side=LEFT)
        Label(frame_title, text=f"{chucVu}", font=(
            "Arial", 13), bg="#f0f4f8").pack(side=LEFT)

        today = datetime.now()
        nam_hien_tai = today.year
        ds_nam_hoat_dong = self.load_years()

        self.cbo_Thang = ttk.Combobox(self.root, width=10, font=(
            "Arial", 13), state="normal", values=[f"Tháng {i}" for i in range(1, 13)])
        self.cbo_Thang.current(today.month - 1)
        self.cbo_Thang.grid(row=0, column=1, padx=10)

        self.cbo_Nam = ttk.Combobox(self.root, width=10, font=(
            "Arial", 13), state="normal", values=[f"Năm {i}" for i in ds_nam_hoat_dong])
        self.cbo_Nam.current(ds_nam_hoat_dong.index(nam_hien_tai))
        self.cbo_Nam.grid(row=0, column=2, padx=5)

        def loc_table():
            self.refresh_table()

        self.btn_filter = Button(
            self.root, text="Lọc", font=("Arial", 13), bg="#3D2AEE", width=5, fg="white", command=loc_table)
        self.btn_filter.grid(row=0, column=3, padx=5)

        self.create_frame_table()
        self.create_frame_duoi_table()
        self.refresh_table()

    def get_ten_nhan_vien_va_luong_co_ban_va_chuc_vu(self, maNhanVien):
        try:
            with open(self.file_account, "r", encoding="utf-8") as f:
                dataAccount = json.load(f)
            for nv in dataAccount:
                if nv["tenTaiKhoan"] == maNhanVien:
                    return nv["tenNguoiDung"], nv["luongCoBan"], nv["loaiTaiKhoan"]

            with open(self.file_driver, "r", encoding="utf-8") as f:
                dataDriver = json.load(f)
            for tx in dataDriver:
                if tx["maTaiXe"] == maNhanVien:
                    return (tx["tenTaiXe"], tx["luongCoBan"], "Nhân viên giao hàng")

            return "Không tìm thấy"
        except Exception as e:
            mb.showerror("Lỗi", f"Lỗi đọc file driver: {e}")
            return "Không xác định"

    def load_years(self):
        try:
            with open(self.fileyear, "r", encoding="utf-8") as file:
                years = json.load(file)
        except (FileNotFoundError, json.JSONDecodeError):
            years = []

        current_year = datetime.now().year
        if current_year not in years:
            years.append(current_year)
            years.sort()

            with open(self.fileyear, "w", encoding="utf-8") as file:
                json.dump(years, file, indent=4)

        return years

    def create_frame_table(self):
        self.frame_table = Frame(self.root)
        self.frame_table.grid(row=1, column=0, sticky="news", columnspan=4)

        columns = ("Ngày", "Thứ", "Số giờ làm việc")
        self.table = ttk.Treeview(self.frame_table, columns=columns,
                                  show="headings", height=15)
        for col in columns:
            self.table.heading(col, text=col)
            self.table.column(col, width=150)

        self.table.tag_configure("oddrow", background="#f0f0f0")
        self.table.tag_configure("evenrow", background="#ffffff")
        self.scrollbar = ttk.Scrollbar(
            self.frame_table, orient="vertical", command=self.table.yview)
        self.table.configure(yscroll=self.scrollbar.set)
        self.scrollbar.pack(side="right", fill="y")
        self.table.pack(fill="both", expand=True)

        # self.table.bind("<<TreeviewSelect>>", lambda e: self.row_selection())

    def create_frame_duoi_table(self):
        self.frame_duoi_table = Frame(self.root)
        self.frame_duoi_table.grid(
            row=2, column=0, sticky="news", columnspan=4)

        self.frame_duoi_table.columnconfigure(0, weight=1)
        self.frame_duoi_table.columnconfigure(1, weight=10)

        self.frame_left = Frame(self.frame_duoi_table)
        self.frame_left.grid(row=0, column=0, sticky="news")

        self.create_frame_buttons()
        self.create_frame_sl_gio_lam()

    def create_frame_buttons(self):
        def nhap_excel():
            file_path = fd.askopenfilename(
                filetypes=[("Excel files", "*.xlsx")])
            if not file_path:
                return

            try:
                wb = openpyxl.load_workbook(file_path)
                if self.maNhanVien not in wb.sheetnames:
                    mb.showerror(
                        "Lỗi", f"Không tìm thấy sheet '{self.maNhanVien}' trong file Excel.")
                    return

                ws = wb[self.maNhanVien]
                lich_lam_viec = []

                for row in ws.iter_rows(min_row=8, values_only=True):
                    ngay_str, thu, gioLam = row
                    if ngay_str is None or thu is None:
                        continue

                    if isinstance(ngay_str, datetime):
                        ngay = ngay_str
                    elif isinstance(ngay_str, date):
                        ngay = datetime.combine(ngay_str, datetime.min.time())
                    elif isinstance(ngay_str, str):
                        try:
                            ngay = datetime.strptime(
                                ngay_str.strip(), "%d/%m/%Y")
                        except ValueError:
                            mb.showerror(
                                "Lỗi định dạng", f"Ngày không đúng định dạng dd/mm/yyyy: {ngay_str}")
                            continue
                    else:
                        mb.showerror(
                            "Lỗi dữ liệu", f"Không xác định kiểu dữ liệu của ngày: {ngay_str}")
                        continue

                    self.thang = ngay.month
                    self.nam = ngay.year

                    try:
                        if gioLam is None:
                            int_gio_lam = 0
                        elif isinstance(gioLam, str):
                            int_gio_lam = int(
                                gioLam.strip()) if gioLam.strip().isdigit() else 0
                        else:
                            int_gio_lam = int(gioLam)
                    except ValueError:
                        int_gio_lam = 0

                    lich_lam_viec.append({
                        "ngay": ngay.day,
                        "thu": thu,
                        "gioLam": int_gio_lam
                    })

                with open(self.filename, "r", encoding="utf-8") as f:
                    data = json.load(f)

                found = False
                for item in data:
                    if (item["maNhanVien"] == self.maNhanVien and
                        item["Thang"] == self.thang and
                            item["Nam"] == self.nam):
                        item["lichLamViec"] = lich_lam_viec  # Ghi đè
                        found = True
                        break

                if not found:
                    data.append({
                        "maNhanVien": self.maNhanVien,
                        "Thang": self.thang,
                        "Nam": self.nam,
                        "lichLamViec": lich_lam_viec
                    })

                with open(self.filename, "w", encoding="utf-8") as file:
                    json.dump(data, file, ensure_ascii=False, indent=4)

                mb.showinfo("Thành công", "Nhập dữ liệu thành công!")
                self.dataLich = self.load_detail_employee(self.filename)
                self.refresh_table()

            except Exception as e:
                mb.showerror("Lỗi", f"Đã xảy ra lỗi khi đọc file:\n{e}")

        self.frame_buttons = Frame(self.frame_left)
        self.frame_buttons.grid(row=0, column=0, sticky="news")

        self.btn_nhap_excel = Button(self.frame_buttons, text="Nhập file Excel", width=15, height=2, font=(
            "Arial", 13, "bold"), bg="#00B050", fg="white", state="normal", command=nhap_excel)
        self.btn_nhap_excel.grid(row=0, column=0, padx=10, pady=5)

    def create_frame_sl_gio_lam(self):
        self.frame_tong_gio_lam = Frame(self.frame_duoi_table)
        self.frame_tong_gio_lam.grid(row=0, column=1, sticky="news")
        self.frame_tong_gio_lam.columnconfigure(0, weight=1)

        self.tong_gio_lam = 0

        self.lbl_hien_thi_tong_gio_lam = Label(
            self.frame_tong_gio_lam, text=f"Tổng số giờ làm: {self.tong_gio_lam}", anchor="e", font=("Arial", 13), bg="#f0f4f8")
        self.lbl_hien_thi_tong_gio_lam.grid(
            row=0, column=0, sticky="e", padx=20, pady=5)

    def refresh_table(self):
        for row in self.table.get_children():
            self.table.delete(row)

        thang = int(self.cbo_Thang.get().replace("Tháng ", ""))
        nam = int(self.cbo_Nam.get().replace("Năm ", ""))

        lich_day_du = self.get_schedule_for_employee(
            self.maNhanVien, self.dataLich)

        lich_loc = []

        for item in lich_day_du:
            d = datetime.strptime(item.ngay, "%d/%m/%Y")
            if d.month == thang and d.year == nam:
                lich_loc.append(item)

        if not lich_loc:
            so_ngay = calendar.monthrange(nam, thang)[1]

            for ngay in range(1, so_ngay + 1):
                d = datetime(nam, thang, ngay)
                thu = d.strftime("%A")
                thu_vn = {
                    "Monday": "Hai", "Tuesday": "Ba", "Wednesday": "Tư",
                    "Thursday": "Năm", "Friday": "Sáu",
                    "Saturday": "Bảy", "Sunday": "Chủ nhật"
                }.get(thu, thu)

                lich_loc.append(LichLamViec(
                    ngay=d.strftime("%d/%m/%Y"),
                    thu=thu_vn,
                    gioLam=0
                ))
        self.lich_Lam_Viec = lich_loc

        self.tong_gio_lam = 0

        for index, item in enumerate(self.lich_Lam_Viec):
            tag = "evenrow" if index % 2 == 0 else "oddrow"
            self.table.insert("", END, values=(
                item.ngay, item.thu, item.gioLam), tags=(tag,))
            self.tong_gio_lam += item.gioLam

        self.lbl_hien_thi_tong_gio_lam.config(
            text=f"Tổng số giờ làm: {self.tong_gio_lam}")

    def load_detail_employee(self, filename):
        try:
            with open(filename, "r", encoding="utf-8") as file:
                data = json.load(file)
            result = {}
            for item in data:
                ma = item["maNhanVien"]
                nam = item["Nam"]
                thang = item["Thang"]
                lich_raw = item["lichLamViec"]

                lich = []
                for x in lich_raw:
                    ngay_str = f"{x['ngay']:02d}/{thang:02d}/{nam}"
                    lich.append(LichLamViec(ngay=ngay_str,
                                thu=x["thu"], gioLam=x["gioLam"]))

                if ma in result:
                    result[ma].extend(lich)
                else:
                    result[ma] = lich

            return result
        except (json.JSONDecodeError, FileNotFoundError):
            return []

    def get_schedule_for_employee(self, maNhanVien, data):
        return data.get(maNhanVien, [])
