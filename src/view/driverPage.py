from tkinter import *
from tkinter import messagebox as mb
from tkinter import ttk
from tkinter import filedialog as fd
import json
from src.Model.TaiXe import TaiXe
from src.Model.NhanVien import NhanVien
import random
import re
import openpyxl
from datetime import datetime, date


class DriverPage:
    def __init__(self, root, switch_page):
        self.root = root
        self.root.configure(bg="#f0f4f8")
        self.root.columnconfigure(0, weight=1)

        self.switch_page = switch_page
        self.filename = "src/database/driver.json"
        self.fileAccount = "src/database/account.json"
        self.dstx = []
        self.data = self.load_data_from_file(self.filename)
        self.dataAccount = self.load_data_from_file(self.fileAccount)

        self.dstx = [TaiXe.from_dict(taixe) for taixe in self.data]
        self.dstk = [NhanVien.from_dict(user) for user in self.dataAccount]

        self.treeview_bind_id = None

        self.create_frame_account_table()
        self.create_frame_duoi_table()

    def row_selection(self):
        selected_item = self.accountTable.focus()
        if selected_item:
            item_values = self.accountTable.item(selected_item)['values']
            if item_values:
                self.btn_remove.config(state="normal")
                self.btn_edit.config(state="normal")

                for e in self.entries:
                    e.config(state="normal")
                    e.delete(0, END)

                self.entry_ma.insert(0, item_values[0])
                self.entry_ten.insert(0, item_values[1])
                self.entry_sdt.insert(0, str(item_values[2]).zfill(10))
                self.entry_diachi.insert(0, str(item_values[3]))
                self.entry_cccd.insert(
                    0, str(item_values[4]).zfill(12))
                self.entry_bank.insert(0, item_values[5])
                self.entry_stk.insert(0, item_values[6])

                luong = str(item_values[7]).replace(".", "").replace("đ", "")
                self.entry_lcb.insert(0, luong)

                for e in self.entries:
                    e.config(state="readonly")

        self.maTaiXe = self.entry_ma.get().strip()

    def create_frame_account_table(self):
        def search():
            for item in self.accountTable.get_children():
                self.accountTable.delete(item)
            nd_timkiem = self.entry_tim_kiem.get().strip()
            tx_timkiem = []
            for tx in self.data:
                if nd_timkiem != "":
                    if str(tx["maTaiXe"]) == nd_timkiem or str(tx["tenTaiXe"]) == nd_timkiem or str(tx["soDienThoai"]) == nd_timkiem or str(tx["cccd"]) == nd_timkiem or str(tx["nganHang"]) == nd_timkiem or str(tx["stk_nganHang"]) == nd_timkiem:
                        tx_timkiem.append(tx)

                else:
                    tx_timkiem.append(tx)

            for index, item in enumerate(tx_timkiem):
                tag = "evenrow" if index % 2 == 0 else "oddrow"
                self.accountTable.insert("", END, values=(
                    item["maTaiXe"], item["tenTaiXe"], str(
                        item["soDienThoai"]),
                    str(item["diaChi"]), str(item["cccd"]), item["nganHang"], str(item["stk_nganHang"]), "{:,}đ".format(int(item["luongCoBan"])).replace(",", ".")), tags=(tag,))

        self.frame_search = Frame(self.root)
        self.frame_search.grid(row=0, column=0, sticky="e", padx=50, pady=5)

        self.entry_tim_kiem = Entry(
            self.frame_search, width=20, font=("Arial", 13))
        self.entry_tim_kiem.grid(row=0, column=0, padx=(0, 20))

        self.btn_search = Button(
            self.frame_search, width=10, text="Tìm kiếm", font=("Arial", 13, "bold"),
            bg="black", fg="white", bd=1, command=search
        )
        self.btn_search.grid(row=0, column=1)

        self.frame_table = Frame(self.root)
        self.frame_table.grid(row=1, column=0, sticky="news")

        columns = ("#1", "#2", "#3", "#4", "#5", "#6", "#7", "#8")
        self.accountTable = ttk.Treeview(
            self.frame_table, columns=columns, show="headings", height=7)
        self.accountTable.heading("#1", text="Mã tài xế")
        self.accountTable.heading("#2", text="Tên tài xế")
        self.accountTable.heading("#3", text="Số điện thoại")

        self.accountTable.heading("#4", text="Địa chỉ")
        self.accountTable.heading("#5", text="CCCD")
        self.accountTable.heading("#6", text="Ngân hàng")
        self.accountTable.heading("#7", text="Số tài khoản")
        self.accountTable.heading("#8", text="Lương cơ bản")

        self.accountTable.column("#1", width=30, anchor="center")
        self.accountTable.column("#2", width=50, anchor="center")
        self.accountTable.column("#3", width=50, anchor="center")
        self.accountTable.column("#4", width=50, anchor="center")
        self.accountTable.column("#5", width=30, anchor="center")
        self.accountTable.column("#6", width=50, anchor="center")
        self.accountTable.column("#7", width=50, anchor="center")
        self.accountTable.column("#8", width=50, anchor="center")

        self.accountTable.tag_configure("oddrow", background="#f0f0f0")
        self.accountTable.tag_configure("evenrow", background="#ffffff")

        # Thêm scrollbar dọc
        self.scrollbar = ttk.Scrollbar(
            self.frame_table, orient="vertical", command=self.accountTable.yview)
        self.accountTable.configure(yscroll=self.scrollbar.set)
        self.scrollbar.pack(side="right", fill="y")
        self.accountTable.pack(fill="both", expand=True)

        for index, item in enumerate(self.data):
            tag = "evenrow" if index % 2 == 0 else "oddrow"
            self.accountTable.insert("", END, values=(
                item["maTaiXe"], item["tenTaiXe"], str(item["soDienThoai"]),
                str(item["diaChi"]), str(item["cccd"]), item["nganHang"], str(item["stk_nganHang"]), "{:,}đ".format(int(item["luongCoBan"])).replace(",", ".")), tags=(tag,))

        self.treeview_bind_id = self.accountTable.bind(
            "<<TreeviewSelect>>", lambda e: self.row_selection())

    def create_frame_duoi_table(self):
        self.frame_duoi_table = Frame(self.root)
        self.frame_duoi_table.grid(row=2, column=0, sticky="news", pady=10)

        self.frame_duoi_table.columnconfigure(0, weight=3)
        self.frame_duoi_table.columnconfigure(1, weight=1)
        self.frame_duoi_table.rowconfigure(0, weight=1)

        self.frame_left = Frame(self.frame_duoi_table)
        self.frame_left.grid(row=0, column=0, sticky="nsew")

        self.frame_left.columnconfigure(0, weight=1)
        self.frame_left.rowconfigure(0, weight=0)
        self.frame_left.rowconfigure(1, weight=1)

        self.create_frame_button()
        self.create_entry()

        self.create_frame_hien_thi_so_luong()

    def create_frame_button(self):
        def refresh_entry():
            for e in self.entries:
                e.config(state="normal")
                e.delete(0, END)

            for e in self.entries:
                e.config(state="readonly")

        def refresh_table():
            for item in self.accountTable.get_children():
                self.accountTable.delete(item)

            for index, tx in enumerate(self.dstx):
                tag = "evenrow" if index % 2 == 0 else "oddrow"
                self.accountTable.insert("", END, values=(
                    tx.maTaiXe, tx.tenTaiXe, str(
                        tx.soDienThoai),
                    str(tx.diaChi), str(tx.cccd), tx.nganHang, str(tx.stk_nganHang), "{:,}đ".format(int(tx.luongCoBan)).replace(",", ".")), tags=(tag,))

                self.lbl_so_luong_taixe.config(
                    text=f"Tổng số tài xế: {len(self.dstx)}")

            refresh_entry()

        def add_user():
            if self.btn_add["text"] == "Thêm":
                for e in self.entries:
                    e.config(state="normal")
                    e.delete(0, END)

                self.btn_add.config(text="Xác nhận")
                self.btn_remove.config(state="disabled")
                self.btn_edit.config(state="disabled")
                self.btn_refresh.config(state="disabled")
                self.btn_nhap_excel.config(state="disabled")
                self.btn_xuat_excel.config(state="disabled")
                self.btn_cancel.config(state="normal")

                def generating_random_id():
                    # Kiểu dữ liệu Set (tập hợp) -> Không chứa phần tử trùng
                    existings_ids = {driver["maTaiXe"] for driver in self.data}

                    while True:
                        random_id = "TX" + \
                            "".join([str(random.randint(0, 9))
                                    for _ in range(10)])
                        if random_id not in existings_ids:
                            return random_id

                new_id = generating_random_id()
                self.entry_ma.insert(0, new_id)
                self.entry_ma.config(state="readonly")

                self.treeview_bind_id = self.accountTable.unbind(
                    "<<TreeviewSelect>>")

            elif self.btn_add["text"] == "Xác nhận":
                ma = self.entry_ma.get()
                ten = self.entry_ten.get().strip()
                sdt = self.entry_sdt.get().strip()
                diaChi = self.entry_diachi.get().strip()
                cccd = self.entry_cccd.get().strip()
                nganHang = self.entry_bank.get().strip()
                stk = self.entry_stk.get().strip()
                lcb_raw = self.entry_lcb.get().strip()

                if ten == "":
                    mb.showwarning("Lỗi", "Vui lòng nhập họ tên!")
                    return
                if sdt == "":
                    mb.showwarning("Lỗi", "Vui lòng nhập số điện thoại!")
                    return

                if cccd == "":
                    mb.showwarning("Lỗi", "Vui lòng nhập CCCD!")
                    return

                try:
                    lcb = int(lcb_raw) if lcb_raw != "" else 0
                except ValueError:
                    mb.showwarning(
                        "Lỗi", "Lương cơ bản không hợp lệ! Vui lòng nhập số.")
                    return

                if len(sdt) != 10:
                    mb.showwarning(
                        "Lỗi", "Vui lòng nhập đúng định dạng số điện thoại (10 chữ số)!")
                    return

                if not sdt.startswith(("09", "03", "08", "07", "05")):
                    mb.showwarning(
                        "Lỗi", "Số điện thoại không hợp lệ! Vui lòng nhập lại.")
                    return

                if len(cccd) != 12:
                    mb.showwarning(
                        "Lỗi", "Vui lòng nhập đúng định dạng CCCD (12 chữ số)!")
                    return

                for tx in self.dstx:
                    if cccd == tx.cccd:
                        mb.showwarning(
                            "Lỗi", "Đã tồn tại nhân viên có CCCD này!")
                        return

                for user in self.dstk:
                    if cccd == user.cccd:
                        mb.showwarning(
                            "Lỗi", "Đã tồn tại nhân viên có CCCD này!")
                        return

                for user in self.dstk:
                    if sdt == user.so_dien_thoai:
                        mb.showwarning(
                            "Lỗi", "Đã tồn tại nhân viên có số điện thoại này!")
                        return

                for tx in self.dstx:
                    if sdt == tx.soDienThoai:
                        mb.showwarning(
                            "Lỗi", "Đã tồn tại nhân viên có số điện thoại này!")
                        return

                taiXe = TaiXe(ma, ten, sdt, str(diaChi),
                              cccd, nganHang, stk, int(lcb))
                self.dstx.append(taiXe)
                self.save_data(self.filename)
                mb.showinfo(
                    "Thông báo", f"Thêm thông tin của nhân viên {ten} thành công.")

                refresh_table()

                self.btn_add.config(text="Thêm")
                self.btn_refresh.config(state="normal")
                self.btn_nhap_excel.config(state="normal")
                self.btn_xuat_excel.config(state="normal")
                self.btn_cancel.config(state="disabled")

                self.accountTable.bind(
                    "<<TreeviewSelect>>", lambda e: self.row_selection())

        def remove_user():
            ma = self.entry_ma.get()
            answer = mb.askyesno(
                "Thông báo", f"Bạn có chắc chắn muốn xóa nhân viên mã {ma} không?")
            if answer:
                for tx in self.dstx:
                    if tx.maTaiXe == ma:
                        self.dstx.remove(tx)
                        self.save_data(self.filename)
                        refresh_table()

                        mb.showinfo(
                            "Thông báo", f"Đã xóa thành công nhân viên mã {ma}.")

                        self.btn_remove.config(state="disabled")
                        self.btn_edit.config(state="disabled")
                        return
            return

        def edit():
            ma = self.entry_ma.get()

            if self.btn_edit["text"] == "Sửa":
                self.btn_add.config(state="disabled")
                self.btn_refresh.config(state="disabled")
                self.btn_remove.config(state="disabled")
                self.btn_nhap_excel.config(state="disabled")
                self.btn_xuat_excel.config(state="disabled")

                self.btn_edit.config(text="Xác nhận")
                self.btn_cancel.config(state="normal")

                for e in self.entries:
                    if e not in [self.entry_ma]:
                        e.config(state="normal")

                self.treeview_bind_id = self.accountTable.unbind(
                    "<<TreeviewSelect>>")

            else:
                ten_moi = self.entry_ten.get().strip()
                sdt_moi = self.entry_sdt.get().strip()
                diaChi_moi = self.entry_diachi.get().strip()
                cccd_moi = self.entry_cccd.get().strip()
                nganHang_moi = self.entry_bank.get().strip()
                stk_moi = self.entry_stk.get().strip()
                new_lcb_raw = self.entry_lcb.get().strip()

                if ten_moi == "":
                    mb.showwarning("Lỗi", "Vui lòng nhập họ tên!")
                    return

                if sdt_moi == "":
                    mb.showwarning("Lỗi", "Vui lòng nhập số điện thoại!")
                    return

                if cccd_moi == "":
                    mb.showwarning("Lỗi", "Vui lòng nhập CCCD!")
                    return

                if len(cccd_moi) != 12:
                    mb.showwarning(
                        "Lỗi", "Vui lòng nhập đúng định dạng CCCD (12 chữ số)!")
                    return

                for user in self.dstk:
                    if user.ten_tai_khoan != ma and user.cccd == cccd_moi:
                        return mb.showwarning("Lỗi", "Đã tồn tại nhân viên có CCCD này!")

                for tx in self.dstx:
                    if tx.maTaiXe != ma and tx.cccd == cccd_moi:
                        return mb.showwarning("Lỗi", "Đã tồn tại nhân viên có CCCD này!")

                if len(sdt_moi) != 10:
                    mb.showwarning(
                        "Lỗi", "Vui lòng nhập đúng định dạng số điện thoại (10 chữ số)!")
                    return

                if not sdt_moi.startswith(("09", "03", "08", "07", "05")):
                    mb.showwarning(
                        "Lỗi", "Số điện thoại không hợp lệ! Vui lòng nhập lại.")
                    return

                for user in self.dstk:
                    if user.ten_tai_khoan != ma and user.so_dien_thoai == sdt_moi:
                        return mb.showwarning("Lỗi", "Đã tồn tại nhân viên có số điện thoại này!")
                for tx in self.dstx:
                    if tx.maTaiXe != ma and tx.soDienThoai == sdt_moi:
                        return mb.showwarning("Lỗi", "Đã tồn tại nhân viên có số điện thoại này!")

                try:
                    lcb_moi = int(new_lcb_raw) if new_lcb_raw != "" else 0
                except ValueError:
                    mb.showwarning(
                        "Lỗi", "Lương cơ bản không hợp lệ! Vui lòng nhập số.")
                    return

                answer = mb.askyesno(
                    "Thông báo", f"Bạn có chắc muốn đổi thông tin cho nhân viên mã {ma} không?")
                if answer:
                    for tx in self.dstx:
                        if tx.maTaiXe == ma:
                            tx.tenTaiXe = ten_moi
                            tx.soDienThoai = sdt_moi
                            tx.diaChi = str(diaChi_moi)
                            tx.cccd = cccd_moi
                            tx.nganHang = nganHang_moi
                            tx.stk_nganHang = stk_moi
                            tx.luongCoBan = int(lcb_moi)

                            self.save_data(self.filename)
                            refresh_table()

                            mb.showinfo(
                                "Thông báo", f"Đã thay đổi thông tin thành công cho nhân viên mã {ma}")

                            self.btn_add.config(state="normal")
                            self.btn_refresh.config(state="normal")
                            self.btn_edit.config(text="Sửa")
                            self.btn_edit.config(state="disabled")
                            self.btn_cancel.config(state="disabled")
                            self.btn_nhap_excel.config(state="normal")
                            self.btn_xuat_excel.config(state="normal")

                            self.accountTable.bind(
                                "<<TreeviewSelect>>", lambda e: self.row_selection())
                            return

        def cancel():
            if self.btn_add["text"] == "Xác nhận":
                answer = mb.askyesno(
                    "Thông báo", "Bạn có chắc muốn hủy thao tác thêm không?")

                if answer:
                    self.btn_add.config(text="Thêm")
                    self.btn_refresh.config(state="normal")
                    self.btn_cancel.config(state="disabled")
                    self.btn_nhap_excel.config(state="normal")
                    self.btn_xuat_excel.config(state="normal")

                    refresh_entry()

                    self.accountTable.bind(
                        "<<TreeviewSelect>>", lambda e: self.row_selection())
                    return
                return

            elif self.btn_edit["text"] == "Xác nhận":
                answer = mb.askyesno(
                    "Thông báo", "Bạn có chắc muốn hủy thao tác sửa không?")

                if answer:
                    self.btn_edit.config(text="Sửa")
                    self.btn_edit.config(state="disabled")
                    self.btn_add.config(state="normal")
                    self.btn_remove.config(state="disabled")
                    self.btn_refresh.config(state="normal")
                    self.btn_nhap_excel.config(state="normal")
                    self.btn_xuat_excel.config(state="normal")

                    self.btn_cancel.config(state="disabled")
                    refresh_entry()

                    self.accountTable.bind(
                        "<<TreeviewSelect>>", lambda e: self.row_selection())
                    return
                return

        def xuat_excel():
            if not self.dstx:
                mb.showwarning("Thông báo", "Không có dữ liệu để xuất!")
                return

            file_path = fd.asksaveasfilename(defaultextension=".xlsx", filetypes=[
                                             ("Excel files", "*.xlsx")], initialfile="Danh_sach_shipper.xlsx")
            if not file_path:
                return

            wb = openpyxl.Workbook()
            ws = wb.active  # Worksheet
            ws.title = "Danh sách tài xế"

            headers = ["Mã tài xế", "Tên  tài xế", "Số điện thoại",
                       "Địa chỉ", "CCCD", "Ngân hàng", "STK ngân hàng", "Lương cơ bản"]
            ws.append(headers)

            for tx in self.dstx:
                row = [tx.maTaiXe, tx.tenTaiXe, str(tx.soDienThoai), str(tx.diaChi), str(
                    tx.cccd), tx.nganHang, str(tx.stk_nganHang), "{:,}đ".format(int(tx.luongCoBan)).replace(",", ".")]

                ws.append(row)

            try:
                wb.save(file_path)
                mb.showinfo("Thành công", "Xuất file Excel thành công!")
            except Exception as e:
                mb.showerror("Lỗi", f"Không thể lưu file:\n{e}")

        def nhap_excel():
            file_path = fd.askopenfilename(
                filetypes=[("Excel files", "*.xlsx")])
            if not file_path:
                return

            try:
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active

                errors = []

                # bỏ dòng tiêu đề, chỉ lấy từ dòng 2 trở đi.
                rows = list(ws.iter_rows(min_row=2, values_only=True))

                for idx, row in enumerate(rows, start=2):
                    ma, ten, sdt, diaChi, cccd, nganHang, stk, lcb = [
                        str(cell or "") for cell in row]

                    luong = lcb.replace("đ", "").replace(
                        ".", "") if lcb else "0"

                    row_errors = []

                    if not ma:
                        row_errors.append("Mã tài xế không được bỏ trống.")
                    else:
                        ma_conflict = any(tx.maTaiXe ==
                                          ma for tx in self.dstx)
                        if ma_conflict:
                            row_errors.append(
                                f"Mã tài xế '{ma}' đã tồn tại ở nhân viên khác.")

                    # Kiểm tra định dạng SĐT
                    if not sdt.isdigit() or len(sdt) != 10:
                        row_errors.append(
                            f"SĐT '{sdt}' không hợp lệ (phải đủ 10 chữ số).")

                    if not cccd.isdigit() or len(cccd) != 12:
                        row_errors.append(
                            f"CCCD '{cccd}' không hợp lệ (phải đủ 12 chữ số).")

                    # Kiểm tra trùng CCCD
                    cccd_conflict = any(
                        (tx.cccd == cccd and tx.maTaiXe != ma) for tx in self.dstx
                    ) or any(
                        (user.cccd == cccd and user.ten_tai_khoan != ma) for user in self.dstk
                    )

                    if cccd_conflict:
                        row_errors.append(
                            f"CCCD '{cccd}' đã tồn tại ở nhân viên khác.")

                    # Kiểm tra trùng SĐT
                    sdt_conflict = any(
                        (tx.soDienThoai == sdt and tx.maTaiXe != ma) for tx in self.dstx
                    ) or any(
                        (user.so_dien_thoai == sdt and user.ten_tai_khoan != ma) for user in self.dstk
                    )

                    if sdt_conflict:
                        row_errors.append(
                            f"SĐT '{sdt}' đã tồn tại ở nhân viên khác.")

                    if row_errors:
                        error_message = f"- Dòng {idx}: " + \
                            "; ".join(row_errors)
                        errors.append(error_message)
                        continue

                    tx_moi = TaiXe(ma, ten, sdt, str(diaChi),
                                   cccd, nganHang, stk, int(luong))

                    found = False
                    for i, tx in enumerate(self.dstx):
                        if tx.maTaiXe == str(ma):
                            self.dstx[i] = tx_moi  # cập nhật tài xế
                            found = True
                            break

                    if not found:
                        self.dstx.append(tx_moi)

                self.save_data(self.filename)
                refresh_table()

                if errors:
                    mb.showwarning(
                        "Một số dòng bị lỗi",
                        "\n".join(errors)
                    )
                else:
                    mb.showinfo(
                        "Thành công", "Nhập dữ liệu từ Excel thành công!")

            except Exception as e:
                mb.showerror("Lỗi", f"Không thể đọc file:\n{e}")

        self.frame_buttons = Frame(self.frame_left)
        self.frame_buttons.grid(row=0, column=0, sticky="news", pady=10)

        self.btn_add = Button(self.frame_buttons, text="Thêm",
                              width=15, height=2, font=("Arial", 13, "bold"), bg="blue", fg="white", command=add_user)
        self.btn_add.grid(row=0, column=0, padx=10)

        self.btn_remove = Button(self.frame_buttons, text="Xóa",
                                 width=15, height=2, font=("Arial", 13, "bold"), bg="#009df7", fg="white", state="disabled", command=remove_user)
        self.btn_remove.grid(row=0, column=1, padx=10)

        self.btn_edit = Button(self.frame_buttons, text="Sửa",
                               width=15, height=2, font=("Arial", 13, "bold"), bg="#00f7f7", fg="white", state="disabled", command=edit)
        self.btn_edit.grid(row=0, column=2, padx=10)

        self.btn_refresh = Button(self.frame_buttons, text="Làm mới", width=15, height=2, font=(
            "Arial", 13, "bold"), bg="#4287f5", fg="white", command=refresh_table)
        self.btn_refresh.grid(row=1, column=0, padx=10, pady=5)

        self.btn_cancel = Button(self.frame_buttons, text="Hủy", width=15, height=2, font=(
            "Arial", 13, "bold"), bg="#f70021", fg="white", state="disabled", command=cancel)
        self.btn_cancel.grid(row=1, column=1, padx=10, pady=5)

        self.btn_xuat_excel = Button(self.frame_buttons, text="Xuất file Excel", width=15, height=2, font=(
            "Arial", 13, "bold"), bg="#00B050", fg="white", state="normal", command=xuat_excel)
        self.btn_xuat_excel.grid(row=1, column=2, padx=10, pady=5)

        self.btn_nhap_excel = Button(self.frame_buttons, text="Nhập file Excel", width=15, height=2, font=(
            "Arial", 13, "bold"), bg="#00B050", fg="white", state="normal", command=nhap_excel)
        self.btn_nhap_excel.grid(row=0, column=3, padx=10, pady=5)

    def create_frame_hien_thi_so_luong(self):
        self.frame_sl_taixe = Frame(self.frame_duoi_table)
        self.frame_sl_taixe.grid(row=0, column=1, sticky="ensw", padx=20)

        tong_so_tx = len(self.data)
        self.lbl_so_luong_taixe = Label(
            self.frame_sl_taixe, text=f"Tổng số tài xế: {tong_so_tx}", font=("Arial", 12), anchor="e")
        self.lbl_so_luong_taixe.grid(row=0, column=0, sticky="ew")

    def create_entry(self):
        def validate_max_length(max_len):
            def validator(value):
                if value == "":
                    return True
                return bool(re.fullmatch(r"\d{0," + str(max_len) + r"}", value))
            return validator

        vcmd_sdt = self.root.register(validate_max_length(10))
        vcmd_cccd = self.root.register(validate_max_length(12))
        vcmd_stk = self.root.register(validate_max_length(20))
        vcmd_lcb = self.root.register(validate_max_length(10))

        self.frame_entries = Frame(self.frame_left)
        self.frame_entries.grid(row=1, column=0, sticky="news", pady=10)

        # Mã tài xế
        self.lbl_ma = Label(
            self.frame_entries, text="Mã tài xế: ", width=15, font=("Arial", 13))
        self.lbl_ma.grid(row=0, column=0, padx=10, pady=8)
        self.entry_ma = Entry(
            self.frame_entries,
            width=20,
            font=("Arial", 13),
            bd=1,
            relief="solid",
            highlightbackground="gray",
            highlightthickness=2,
            highlightcolor="blue",
            state="readonly")  # ko cho người dùng gõ vào
        self.entry_ma.grid(row=0, column=1, padx=10, pady=8)

        # Tên tài xế
        self.lbl_ten = Label(
            self.frame_entries, text="Tên tài xế: ", width=15, font=("Arial", 13))
        self.lbl_ten.grid(row=1, column=0, padx=10, pady=8)
        self.entry_ten = Entry(
            self.frame_entries,
            width=20,
            font=("Arial", 13),
            bd=1,
            relief="solid",
            highlightbackground="gray",
            highlightthickness=2,
            highlightcolor="blue",
            state="readonly")
        self.entry_ten.grid(row=1, column=1, padx=10, pady=8)

        # Số điện thoại
        self.lbl_sdt = Label(
            self.frame_entries, text="Số điện thoại: ", width=15, font=("Arial", 13))
        self.lbl_sdt.grid(row=2, column=0, padx=10, pady=8)
        self.entry_sdt = Entry(
            self.frame_entries,
            width=20,
            font=("Arial", 13),
            bd=1,
            relief="solid",
            highlightbackground="gray",
            highlightthickness=2,
            highlightcolor="blue",
            state="readonly",
            validate="key",
            validatecommand=(vcmd_sdt, "%P"),
        )
        self.entry_sdt.grid(row=2, column=1, padx=10, pady=8)

        # Địa chỉ
        self.lbl_diachi = Label(
            self.frame_entries, text="Địa chỉ: ", width=15, font=("Arial", 13))
        self.lbl_diachi.grid(row=0, column=2, padx=10, pady=8)
        self.entry_diachi = Entry(
            self.frame_entries,
            width=20,
            font=("Arial", 13),
            bd=1,
            relief="solid",
            highlightbackground="gray",
            highlightthickness=2,
            highlightcolor="blue",
            state="readonly")
        self.entry_diachi.grid(row=0, column=3, padx=10, pady=8)

        # CMND - CCCD
        self.lbl_cccd = Label(
            self.frame_entries, text="CCCD: ", width=15, font=("Arial", 13))
        self.lbl_cccd.grid(row=1, column=2, padx=10, pady=8)
        self.entry_cccd = Entry(
            self.frame_entries,
            width=20,
            font=("Arial", 13),
            bd=1,
            relief="solid",
            highlightbackground="gray",
            highlightthickness=2,
            highlightcolor="blue",
            state="readonly",
            validate="key",
            validatecommand=(vcmd_cccd, "%P"))
        self.entry_cccd.grid(row=1, column=3, padx=10, pady=8)

        # Ngân hàng
        self.lbl_bank = Label(
            self.frame_entries, text="Ngân hàng: ", width=15, font=("Arial", 13))
        self.lbl_bank.grid(row=2, column=2, padx=10, pady=8)
        self.entry_bank = Entry(
            self.frame_entries,
            width=20,
            font=("Arial", 13),
            bd=1,
            relief="solid",
            highlightbackground="gray",
            highlightthickness=2,
            highlightcolor="blue",
            state="readonly")
        self.entry_bank.grid(row=2, column=3, padx=10, pady=8)

        # STK - Bank
        self.lbl_stk = Label(
            self.frame_entries, text="Số tài khoản: ", width=15, font=("Arial", 13))
        self.lbl_stk.grid(row=3, column=2, padx=10, pady=8)
        self.entry_stk = Entry(
            self.frame_entries,
            width=20,
            font=("Arial", 13),
            bd=1,
            relief="solid",
            highlightbackground="gray",
            highlightthickness=2,
            highlightcolor="blue",
            state="readonly",
            validate="key",
            validatecommand=(vcmd_stk, "%P"))
        self.entry_stk.grid(row=3, column=3, padx=10, pady=8)

        # Lương cơ bản
        self.lbl_lcb = Label(
            self.frame_entries, text="Lương cơ bản: ", width=15, font=("Arial", 13))
        self.lbl_lcb.grid(row=3, column=0, padx=10, pady=8)
        self.entry_lcb = Entry(
            self.frame_entries,
            width=20,
            font=("Arial", 13),
            bd=1,
            relief="solid",
            highlightbackground="gray",
            highlightthickness=2,
            highlightcolor="blue",
            state="readonly",
            validate="key",
            validatecommand=(vcmd_lcb, "%P"))
        self.entry_lcb.grid(row=3, column=1, padx=10, pady=8)

        self.entries = [
            self.entry_ma, self.entry_ten, self.entry_sdt,
            self.entry_diachi, self.entry_cccd, self.entry_bank,
            self.entry_stk, self.entry_lcb
        ]

    def load_data_from_file(self, filename):
        try:
            with open(filename, "r", encoding="utf-8") as file:
                return json.load(file)
        except (json.JSONDecodeError, FileNotFoundError):
            return []

    def save_data(self, filename):
        data = [tx.to_dict() for tx in self.dstx]
        with open(filename, "w", encoding="utf-8") as file:
            json.dump(data, file, ensure_ascii=False, indent=4)
